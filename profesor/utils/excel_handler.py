import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse
from ejercicios.models import Ejercicio, TipoEjercicio
from core.models import Unidad, Materia
from django.db import transaction
from datetime import datetime


class ExcelEjerciciosHandler:
    """Maneja la importación y exportación de ejercicios en Excel"""
    
    # Columnas esperadas en el Excel
    COLUMNAS = [
        'enunciado',
        'solucion',
        'dificultad',
        'discriminacion',
        'unidad_id',
        'fuente',
        'licencia',
        'tipos_ejercicio'  # Valores separados por coma: funciones,matrices
    ]
    
    # Opciones válidas
    FUENTES_VALIDAS = ['deepseek', 'llama', 'chatgpt', 'real']
    LICENCIAS_VALIDAS = ['cc-by', 'cc-by-sa', 'cc0', 'mit', 'gpl']
    
    @staticmethod
    def generar_plantilla(materia_id):
        """
        Genera un archivo Excel de plantilla para cargar ejercicios
        CON DROPDOWN DE UNIDADES específicas de la materia
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ejercicios"
        
        # Estilos
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            'enunciado',
            'solucion',
            'dificultad',
            'discriminacion',
            'unidad_id',
            'fuente',
            'licencia',
            'tipos_ejercicio'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Obtener materia y sus unidades
        try:
            materia = Materia.objects.get(id=materia_id)
            unidades = Unidad.objects.filter(materia=materia).order_by('num_unidad')
            
            # ===================================================
            # CREAR DROPDOWN DE UNIDADES
            # ===================================================
            if unidades.exists():
                # Crear hoja oculta con datos de unidades para el dropdown
                ws_data = wb.create_sheet("_UnidadesData")
                
                # Escribir IDs de unidades en la hoja oculta
                for idx, unidad in enumerate(unidades, start=1):
                    ws_data.cell(row=idx, column=1, value=unidad.id)
                
                # Crear la fórmula de validación
                # Rango de datos: desde _UnidadesData!$A$1 hasta la última unidad
                ultima_fila = unidades.count()
                formula = f"_UnidadesData!$A$1:$A${ultima_fila}"
                
                # Crear validación de datos (dropdown)
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=False,
                    showDropDown=True,
                    showErrorMessage=True,
                    errorTitle="ID de Unidad Inválido",
                    error="Por favor selecciona un ID de unidad de la lista"
                )
                
                # Aplicar validación a la columna E (unidad_id) para 1000 filas
                dv.add(f'E2:E1000')
                ws.add_data_validation(dv)
                
                # Ocultar la hoja de datos
                ws_data.sheet_state = 'hidden'
                
                print(f"✅ Dropdown creado con {unidades.count()} unidades")
            
        except Materia.DoesNotExist:
            print("❌ Materia no encontrada")
            unidades = []
        
        # Fila de ejemplo
        ejemplo = [
            'f(x)=5x+4 cuando x=2?',
            '14',
            '0.5',
            '1.0',
            str(unidades.first().id) if unidades.exists() else '1',  # Primer ID disponible
            'real',
            'cc-by',
            'funciones,limites'
        ]
        
        for col_num, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = valor
            cell.border = border
        
        # Hoja de instrucciones
        ws_inst = wb.create_sheet("Instrucciones")
        ws_inst.column_dimensions['A'].width = 50
        ws_inst.column_dimensions['B'].width = 50
        
        instrucciones = [
            ["INSTRUCCIONES PARA CARGAR EJERCICIOS", ""],
            ["", ""],
            ["Campo", "Descripción"],
            ["enunciado", "Texto del problema (máx. 200 caracteres)"],
            ["solucion", "Respuesta correcta (máx. 200 caracteres)"],
            ["dificultad", "Valor entre -3.0 y 3.0"],
            ["discriminacion", "Valor entre 0.01 y 2.0 (default: 1.0)"],
            ["unidad_id", "ID de la unidad - USAR DROPDOWN (ver hoja 'Unidades')"],
            ["fuente", f"Opciones: {', '.join(ExcelEjerciciosHandler.FUENTES_VALIDAS)}"],
            ["licencia", f"Opciones: {', '.join(ExcelEjerciciosHandler.LICENCIAS_VALIDAS)}"],
            ["tipos_ejercicio", "Tipos separados por coma (ver hoja 'Tipos Disponibles')"],
            ["", ""],
            ["IMPORTANTE:", ""],
            ["- No modifiques los nombres de las columnas", ""],
            ["- La fila 2 es un ejemplo, puedes eliminarla", ""],
            ["- Dificultad: valores negativos = fácil, positivos = difícil", ""],
            ["- Discriminación: qué tan bien diferencia el ítem entre estudiantes", ""],
            ["- ⚠️ IMPORTANTE: En 'unidad_id' haz clic y selecciona del dropdown", ""],
            ["- Solo aparecen las unidades de la materia seleccionada", ""],
        ]
        
        for row_num, (campo, desc) in enumerate(instrucciones, 1):
            ws_inst.cell(row=row_num, column=1).value = campo
            ws_inst.cell(row=row_num, column=2).value = desc
            if row_num == 1:
                ws_inst.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="4F46E5")
            elif row_num == 3:
                ws_inst.cell(row=row_num, column=1).font = Font(bold=True)
                ws_inst.cell(row=row_num, column=2).font = Font(bold=True)
            elif "IMPORTANTE" in str(campo):
                ws_inst.cell(row=row_num, column=1).font = Font(bold=True, color="DC2626")
            elif "⚠️" in str(campo):
                ws_inst.cell(row=row_num, column=1).font = Font(bold=True, color="DC2626")
        
        # Hoja de unidades disponibles (VISIBLE)
        try:
            ws_unidades = wb.create_sheet("Unidades Disponibles")
            ws_unidades.column_dimensions['A'].width = 15
            ws_unidades.column_dimensions['B'].width = 10
            ws_unidades.column_dimensions['C'].width = 40
            ws_unidades.column_dimensions['D'].width = 60
            
            # Headers
            headers_unidades = ['ID', 'Número', 'Nombre', 'Objetivo']
            for col_num, header in enumerate(headers_unidades, 1):
                cell = ws_unidades.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos de unidades
            for row_num, unidad in enumerate(unidades, 2):
                ws_unidades.cell(row=row_num, column=1, value=unidad.id).font = Font(bold=True, color="4F46E5")
                ws_unidades.cell(row=row_num, column=2, value=unidad.num_unidad)
                ws_unidades.cell(row=row_num, column=3, value=unidad.nombre)
                ws_unidades.cell(row=row_num, column=4, value=unidad.objetivo)
                
                # Alternar colores de fila
                if row_num % 2 == 0:
                    fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
                    for col in range(1, 5):
                        ws_unidades.cell(row=row_num, column=col).fill = fill
            
            # Nota importante
            nota_row = unidades.count() + 3
            ws_unidades.cell(row=nota_row, column=1, value="⚠️ IMPORTANTE:").font = Font(bold=True, color="DC2626", size=12)
            ws_unidades.cell(row=nota_row + 1, column=1, value="En la hoja 'Ejercicios', la columna 'unidad_id' tiene un dropdown.").font = Font(color="059669")
            ws_unidades.cell(row=nota_row + 2, column=1, value="Haz clic en la celda y selecciona el ID de la unidad deseada.").font = Font(color="059669")
            ws_unidades.cell(row=nota_row + 3, column=1, value="NO escribas el ID manualmente, usa el dropdown.").font = Font(bold=True, color="DC2626")
            
        except Exception as e:
            print(f"Error al cargar unidades: {e}")
        
        # Hoja de tipos disponibles
        ws_tipos = wb.create_sheet("Tipos Disponibles")
        ws_tipos.column_dimensions['A'].width = 20
        ws_tipos.column_dimensions['B'].width = 40
        
        # Headers
        ws_tipos.cell(row=1, column=1, value="Código").font = Font(bold=True)
        ws_tipos.cell(row=1, column=1).fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        ws_tipos.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        
        ws_tipos.cell(row=1, column=2, value="Nombre").font = Font(bold=True)
        ws_tipos.cell(row=1, column=2).fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        ws_tipos.cell(row=1, column=2).font = Font(bold=True, color="FFFFFF")
        
        tipos = TipoEjercicio.objects.all()
        for row_num, tipo in enumerate(tipos, 2):
            ws_tipos.cell(row=row_num, column=1, value=tipo.tipo_ejercicio)
            ws_tipos.cell(row=row_num, column=2, value=tipo.get_tipo_ejercicio_display())
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'plantilla_ejercicios_{materia.nombre.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def validar_fila(row_data, row_num):
        """
        Valida una fila del Excel
        Retorna (es_valida, errores_lista)
        """
        errores = []
        
        # Validar campos obligatorios
        if not row_data.get('enunciado'):
            errores.append(f"Fila {row_num}: 'enunciado' es obligatorio")
        elif len(row_data['enunciado']) > 200:
            errores.append(f"Fila {row_num}: 'enunciado' no puede superar 200 caracteres")
        
        if not row_data.get('solucion'):
            errores.append(f"Fila {row_num}: 'solucion' es obligatoria")
        elif len(row_data['solucion']) > 200:
            errores.append(f"Fila {row_num}: 'solucion' no puede superar 200 caracteres")
        
        # Validar dificultad
        try:
            dificultad = float(row_data.get('dificultad', 0))
            if not (-3.0 <= dificultad <= 3.0):
                errores.append(f"Fila {row_num}: 'dificultad' debe estar entre -3.0 y 3.0")
        except (ValueError, TypeError):
            errores.append(f"Fila {row_num}: 'dificultad' debe ser un número")
        
        # Validar discriminación
        try:
            discriminacion = float(row_data.get('discriminacion', 1.0))
            if not (0.01 <= discriminacion <= 2.0):
                errores.append(f"Fila {row_num}: 'discriminacion' debe estar entre 0.01 y 2.0")
        except (ValueError, TypeError):
            errores.append(f"Fila {row_num}: 'discriminacion' debe ser un número")
        
        # Validar unidad_id
        try:
            unidad_id = int(row_data.get('unidad_id', 0))
            if not Unidad.objects.filter(id=unidad_id).exists():
                errores.append(f"Fila {row_num}: unidad_id={unidad_id} no existe")
        except (ValueError, TypeError):
            errores.append(f"Fila {row_num}: 'unidad_id' debe ser un número entero")
        
        # Validar fuente (opcional)
        fuente = row_data.get('fuente', '').lower()
        if fuente and fuente not in ExcelEjerciciosHandler.FUENTES_VALIDAS:
            errores.append(
                f"Fila {row_num}: 'fuente' debe ser una de: {', '.join(ExcelEjerciciosHandler.FUENTES_VALIDAS)}"
            )
        
        # Validar licencia (opcional)
        licencia = row_data.get('licencia', '').lower()
        if licencia and licencia not in ExcelEjerciciosHandler.LICENCIAS_VALIDAS:
            errores.append(
                f"Fila {row_num}: 'licencia' debe ser una de: {', '.join(ExcelEjerciciosHandler.LICENCIAS_VALIDAS)}"
            )
        
        return len(errores) == 0, errores
    
    @staticmethod
    def procesar_excel(archivo, materia_id, docente=None):
        """
        Procesa un archivo Excel y crea los ejercicios
        Retorna (ejercicios_creados, errores_lista)
        """
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            # Verificar encabezados
            headers = [cell.value for cell in ws[1]]
            if not all(col in headers for col in ExcelEjerciciosHandler.COLUMNAS):
                return 0, ["El archivo no tiene las columnas correctas. Descarga la plantilla."]
            
            ejercicios_creados = 0
            errores = []
            
            # Obtener materia
            try:
                materia = Materia.objects.get(id=materia_id)
            except Materia.DoesNotExist:
                return 0, ["La materia seleccionada no existe"]
            
            # Procesar filas (saltando el header)
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Saltar filas vacías
                    if not any(row):
                        continue
                    
                    # Crear diccionario con los datos
                    row_data = dict(zip(headers, row))
                    
                    # Validar fila
                    es_valida, errores_fila = ExcelEjerciciosHandler.validar_fila(row_data, row_num)
                    
                    if not es_valida:
                        errores.extend(errores_fila)
                        continue
                    
                    # Crear ejercicio
                    try:
                        ejercicio = Ejercicio.objects.create(
                            enunciado=row_data['enunciado'].strip(),
                            solucion=row_data['solucion'].strip(),
                            dificultad=float(row_data['dificultad']),
                            discriminacion=float(row_data.get('discriminacion', 1.0)),
                            unidad_id=int(row_data['unidad_id']),
                            materia=materia,
                            docente=docente,
                            fuente=row_data.get('fuente', '').lower() or None,
                            licencia=row_data.get('licencia', '').lower() or None,
                            is_active=True
                        )
                        
                        # Asociar tipos de ejercicio
                        tipos_str = row_data.get('tipos_ejercicio', '')
                        if tipos_str:
                            tipos_lista = [t.strip().lower() for t in str(tipos_str).split(',')]
                            for tipo_codigo in tipos_lista:
                                try:
                                    tipo_obj = TipoEjercicio.objects.get(tipo_ejercicio=tipo_codigo)
                                    ejercicio.tipo_ejercicio.add(tipo_obj)
                                except TipoEjercicio.DoesNotExist:
                                    errores.append(
                                        f"Fila {row_num}: tipo_ejercicio '{tipo_codigo}' no existe"
                                    )
                        
                        ejercicios_creados += 1
                        
                    except Exception as e:
                        errores.append(f"Fila {row_num}: Error al crear ejercicio - {str(e)}")
            
            return ejercicios_creados, errores
            
        except Exception as e:
            return 0, [f"Error al procesar el archivo: {str(e)}"]